from django.shortcuts import render, redirect, get_object_or_404
from tablib import Dataset
from django.core.exceptions import MultipleObjectsReturned
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import logout, authenticate, login

from .models import *
from client.models import Profile as Profile_client
from .forms import *
from django.urls import reverse, reverse_lazy
from django.views.generic.base import *
from django.views.generic import *
from django.conf import settings
from django.contrib import messages
import pandas as pd
import openpyxl
import os
import math
from os.path import join
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.mail import send_mail
from tablib import Dataset
from transliterate.decorators import transliterate_function
import subprocess
from openpyxl.utils.cell import get_column_letter
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

@transliterate_function(language_code='ru', reversed=True)
def translit(text):
    return text


def main(request, super_us = False):
    if request.user.is_authenticated:
        username = request.user.username
        data = {
            'files': Client.objects.filter(username=username)[::-1],
            'cleared': Cleared.objects.filter(username=username)[::-1],
            'complete': Complete.objects.filter(username=username)[::-1],
                }
        if request.user.is_superuser == 1:
            data['body'] = 'on'
            return render(request, 'base.html', data)
        data['body'] = 'off'

        hol = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        if not os.path.exists(os.path.join(hol, f"media/clients/{translit(username)}")):
            os.mkdir(os.path.join(hol, f"media/clients/{translit(username)}"))

            r = Report.objects.all()[0]
            report_common = r.file.name

            wb = openpyxl.load_workbook(filename=os.path.join(hol, f"media/{report_common}"))

            w = wb.worksheets[0]
            sheet = wb.active
            wb.save(os.path.join(hol, f"media/clients/{translit(username)}/report.xlsx").replace('\\', '/'))
            try:
                try:
                    f = Profile.objects.get(bying_username=username)
                except ObjectDoesNotExist:
                    f = Profile.objects.get(manager_username=username)
                f.report_common = f"media/clients/{translit(username)}/report.xlsx"
                f.save()
            except:
                Profile.objects.create(bying_username=username, manager_username=username, report_common=f"media/clients/{translit(username)}/report.xlsx")
    else:
        return redirect('exel:login')
    return render(request, 'base.html', data)


class RegisterView(CreateView):
    form_class  = RegisterUserForm
    template_name = 'registration/register.html'

    def get_success_url(self):
        return reverse('exel:login')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        return dict(list(context.items()))



class Send_email(TemplateView):
    template_name = "registration/send_email.html"
    def dispatch(self, request, *args, **kwargs):
        send_mail('Тема', 'Тело письма', settings.EMAIL_HOST_USER, ['davydoval2005@gmail.com'])

class Change(TemplateView):
    template_name = "registration/change.html"
    def dispatch(self, request, *args, **kwargs):
        data = {
            }
        return redirect('exel:login')

def Logout(request):
    logout(request)
    return redirect('exel:login')



class Prepare_calc(TemplateView):
    template_name = "prepare_calculation/prepare.html"
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            username = request.user.username
            # In the down def to create a file DMP.xlsx
            path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            data = []

            # CHECK!!! maybe it even doesn't need
            # merge info from profile page
            for user in User.objects.all():
                try:
                    profile = Profile_client.objects.get(user_id=user.id)
                except:
                    continue
                if profile.mediakit != None:
                    if os.path.exists(os.path.join(path, 'sites', translit(user.username), translit(profile.mediakit))):
                        mediakit = os.path.join(path, 'sites', translit(user.username), translit(profile.mediakit))
                    else:
                        mediakit = profile.mediakit_text
                else:
                    mediakit = profile.mediakit_text
                if profile.price != None:
                    if os.path.exists(os.path.join(path, 'sites', translit(user.username), translit(profile.price))):
                        price = os.path.join(path, 'sites', translit(user.username), translit(profile.price))
                    else:
                        price = profile.price_text
                else:
                    price = profile.price_text
                if profile.example != None:
                    if os.path.exists(os.path.join(path, 'sites', translit(user.username), translit(profile.example))):
                        example = os.path.join(path, 'sites', translit(user.username), translit(profile.example))
                    else:
                        example = profile.example_text
                else:
                    example = profile.example_text
                if profile.TT != None:
                    if os.path.exists(os.path.join(path, 'sites', translit(user.username), translit(profile.TT))):
                        TT = os.path.join(path, 'sites', translit(user.username), translit(profile.TT))
                    else:
                        TT = profile.TT_text
                else:
                    TT = profile.TT_text
                try:
                    headers = ['Селлер','Медиакит/прайсы/пример размещения', 'Пример размещения', 'ТТ', 'Контакты', 
                            'AdRiver (пиксель или кликовая)', 'Запуск (нюансы)', 'Доп.аналитика/комментарии', 'Предоплата', 'Преимущества', 'Входной бюджет/мин объем', 'Минусы', 
                            'KPI', 'Название аудитории', 'Описание аудитории', 'Все сезонные коэф. на 2022', 'Сайт', 'Место размещения на сайте и таргетинги', 'Размер (в пикселях) / Формат',
                            'Тип размещения', 'Единица покупки', 'Цена (за единицу покупки), руб.', 'Наценки / Доп. Скидки', 'Скидка, %', 'Частота', 'VTR,%', 'CTR,%',
                            'Ёмкость ~', 'Комментарии']
                    second_part = pd.read_excel(os.path.join(path, 'sites', translit(user.username), 'second_part.xlsx'), engine='openpyxl', header=None, skiprows=1)
                    for row in second_part.values.tolist():
                        data.append([user.username, f'{mediakit}, {price}', example, TT, profile.contacts, profile.AdRiver, profile.launch, 
                        profile.dop_comments, profile.prepayment, profile.advantages, profile.budget, profile.minuses]+row)
                except:
                    continue
                
            #pd.DataFrame(data).to_excel(os.path.join(path, 'sites', 'dmp.xlsx'), index=None)
            df_profile_info = pd.DataFrame(data, dtype="string")
            df_profile_info.columns = headers
            df_buying_info = pd.DataFrame(list(BuyingDB.objects.values_list()), dtype="string")
            df_buying_info.columns = ['#', 'Селлер', 'Сайт', 'скидка', 'Баинговые приоритеты']
            df_buying_info.to_excel(os.path.join(path, 'sites', 'buying.xlsx'), index=None)
            dmp = pd.merge(df_profile_info, df_buying_info, how='left')


            filter = pd.DataFrame()
            column_order = ['Название аудитории', 'Описание аудитории', 'KPI', 'ТТ', 'AdRiver (пиксель или кликовая)', 'Запуск (нюансы)', 
                'скидка', 'Минусы', "Медиакит/прайсы/пример размещения", 'Пример размещения', 
                'Контакты', "Доп.аналитика/комментарии", 'Входной бюджет/мин объем', 'Предоплата', 'Преимущества', 'Баинговые приоритеты', 'Все сезонные коэф. на 2022', 
                'Сайт', 'Место размещения на сайте и таргетинги', 'Размер (в пикселях) / Формат', 'Тип размещения', 'Единица покупки', "Цена (за единицу покупки), руб.",
                'Наценки / Доп. Скидки', 'Скидка, %', 'Частота', 'VTR,%', 'CTR,%', 'Ёмкость ~', 'Комментарии']
            for i in column_order:
                filter[i] = dmp[i]
            len_col = len(filter.index)
            
            filter.insert(loc=filter.columns.get_loc("Единица покупки")+1, column='Период размещения', value=[1]*len_col)
            filter.insert(loc=filter.columns.get_loc("Период размещения")+1, column='', value=['']*len_col)

        
            

            dataclass = {
                'cl': set(Brief.objects.all().values_list('client', flat=True)),
                'n_rk' : set(Brief.objects.all().values_list('name_rk', flat=True)),
                'client': Brief.objects.filter(username=username)[::-1],
                'form': BriefForm,
                #'now' : Brief.objects.filter(username=username),
                }
            if request.method == 'POST':
                discount = request.POST.get('discount', None)
                AK = request.POST.get('AK', None)
                DCM = request.POST.get('DCM', None)
                client = request.POST.get('client', None)
                product = request.POST.get('product', None)
                name_rk = request.POST.get('name_rk', None)
                posad = request.POST.get('posad', None)
                type_act = request.POST.get('type_act', None)
                country = request.POST.get('country', None)
                ca = request.POST.get('ca', None)
                period_c = request.POST.get('period_c', None)
                period_p = request.POST.get('period_p', None)
                KPI = request.POST.get('KPI', None)

                form = BriefForm(request.POST, request.FILES)

                if form.is_valid():
                    ex = request.FILES.get('img')
                    try:
                        brief = Brief.objects.create(user_id=request.user.id,username=username, client=client, product=product,
                                         name_rk=name_rk, posad=posad,
                                         type_act=type_act, country=country, ca=ca,
                                         period_c=period_c, period_p=period_p,
                                         KPI=KPI, img=ex,
                                         discount=discount, AK=AK, DCM=DCM)
                    except (NameError, AttributeError):
                        s = Brief.objects.filter(username=username, client=client)[::-1][0]
                        k = s.img.name
                        brief = Brief.objects.create(username=username, client=client, product=product,
                                         name_rk=name_rk, posad=posad,
                                         type_act=type_act, country=country, ca=ca,
                                         period_c=period_c, period_p=period_p,
                                         KPI=KPI, img=k,
                                         discount=discount, AK=AK, DCM=DCM)
                    except MultipleObjectsReturned:
                        pass
                
                filter = filter.query("`Название аудитории` in [@type_act, 'ВСЕ']")
                filter = filter.query("KPI in [@KPI]")


                comments_margin = filter['Наценки / Доп. Скидки'].tolist()
                filter = filter.drop('Наценки / Доп. Скидки', axis=1)
                
                if not filter.empty:
                    # ОТЧЕТЫ
                    #for report in ReportFile.objects.filter(user_id=request.user.id)[-1]:
                    path_plan = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                    for report in ReportFile.objects.filter(user_id=request.user.id)[::-1]:
                        try:
                            df_report = pd.read_excel(os.path.join(path_plan, report.file.url.replace('/', '\\')[1:]), engine='openpyxl')
                            Client = df_report.apply(lambda row: row.astype(str).str.contains(r'^(?=.*Клиент*)').any(), axis=1)
                            idx_header1 = Client[Client].index[0]
                            Site = df_report.apply(lambda row: row.astype(str).str.contains(r'^(?=.*Сайт*)').any(), axis=1)
                            idx_header2 = Site[Site].index[0]
                            Size = df_report.apply(lambda row: row.astype(str).str.contains(r'^(?=.*Формат*)').any(), axis=1)
                            idx_header3 = Size[Size].index[0]
                            if idx_header1 == idx_header2:
                                df_report = pd.read_excel(os.path.join(path_plan, report.file.url.replace('/', '\\')[1:]), engine='openpyxl', header=idx_header1+1)
                                break
                            elif idx_header1 == idx_header3 or idx_header2 == idx_header3:
                                df_report = pd.read_excel(os.path.join(path_plan, report.file.url.replace('/', '\\')[1:]), engine='openpyxl', header=idx_header3+1)
                                break
                            else:
                                continue
                        except FileNotFoundError:
                            df_report = {}
                            continue

                    len_col = len(filter.index)
                    

                    filter.insert(loc=filter.columns.get_loc("Цена (за единицу покупки), руб.")+1, column='Наценки / Доп. Скидки', value=[1]*len_col)
                    # VTR
                    df_report = pd.DataFrame(df_report)
                    
                    if not df_report.empty:
                        df_report[['Клиент', 'Сайт', 'Размер (в пикселях) / Формат']] = df_report[['Клиент', 'Сайт', 'Размер (в пикселях) / Формат']].astype(str)
                        filter[['VTR,%', 'CTR,%']] = filter[['VTR,%', 'CTR,%']].astype(float)
                        for i in range(len_col):
                            row_vtr = df_report.loc[(df_report['Клиент'].isin([client])) & (df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                            if not row_vtr.empty:
                                filter['VTR,%'].iloc[i] += float(row_vtr[row_vtr.columns[row_vtr.columns.get_loc('VTR,%')+1]].tolist()[0])
                            else:
                                row_vtr = df_report.loc[(df_report['Сегмент'] == type_act) & (df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                                if not row_vtr.empty:
                                    filter['VTR,%'].iloc[i] += float(row_vtr[row_vtr.columns[row_vtr.columns.get_loc('VTR,%')+1]].tolist()[0]) * 0,95
                                else:
                                    row_vtr = df_report.loc[(df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                                    if not row_vtr.empty:
                                        filter['VTR,%'].iloc[i] += float(row_vtr[row_vtr.columns[row_vtr.columns.get_loc('VTR,%')+1]].tolist()[0]) * 0,92
                                    else:
                                        try:
                                            filter['VTR,%'].iloc[i] *= 0,9
                                        except TypeError:
                                            pass

                            row_ctr = df_report.loc[(df_report['Клиент'].isin([client])) & (df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                            if not row_ctr.empty:
                                filter['CTR,%'].iloc[i] += float(row_ctr[row_ctr.columns[row_ctr.columns.get_loc('CTR%')+1]].tolist()[0])
                            else:
                                row_ctr = df_report.loc[(df_report['Сегмент'] == type_act) & (df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                                if not row_ctr.empty:
                                    filter['CTR,%'].iloc[i] += float(row_ctr[row_ctr.columns[row_ctr.columns.get_loc('CTR%')+1]].tolist()[0]) * 0,95
                                else:
                                    row_ctr = df_report.loc[(df_report['Сайт'] == filter['Сайт'].iloc[i]) & (df_report['Размер (в пикселях) / Формат'] == filter['Размер (в пикселях) / Формат'].iloc[i])].tail(1)
                                    if not row_ctr.empty:
                                        filter['CTR,%'].iloc[i] += float(row_ctr[row_ctr.columns[row_ctr.columns.get_loc('CTR%')+1]].tolist()[0]) * 0,92
                                    else:
                                        try:
                                            filter['CTR,%'].iloc[i] *= 0,9
                                        except TypeError:
                                            pass


                    filter.insert(loc=filter.columns.get_loc("Название аудитории"), column='Категория Клиента', value=[type_act]*len_col)
                    filter.insert(loc=filter.columns.get_loc("Запуск (нюансы)")+1, column='коэф. скидки от 1 (min стоимость плана) до  3 (max стоимость плана)', value=[discount]*len_col)

                    # формулы делать вконце, перед эти все столбци должны быть на месте
                    filter.insert(loc=filter.columns.get_loc("Период размещения")+2, column='Общее количество единиц', value=[
                        f'=IF(OR(INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="1000 показов",INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="CPC"),5000,1)' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Общее количество единиц"), column='Количество единиц за период', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+2}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Период размещения")+1}))' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Скидка, %")+1, column='Стоимость размещения после скидки, руб.', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Цена (за единицу покупки), руб.")+1}))*INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+1}))*INDIRECT(ADDRESS({i},{filter.columns.get_loc("Наценки / Доп. Скидки")+1}))-INDIRECT(ADDRESS({i},{filter.columns.get_loc("Скидка, %")+1}))' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Стоимость размещения после скидки, руб.")+1, column='Количество показов', value=[
                        f'=IF(INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="1000 показов",INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+1}))*1000,IF(INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="CPC",INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("CTR,%")+3})),""))' 
                        for i in range(2, len_col+2)]) # +3 в CTR тк добавятся еще Охват технический и Количество просмотров
                    filter.insert(loc=filter.columns.get_loc("Стоимость размещения после скидки, руб."), column='CPM с учетом скидки', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Стоимость размещения после скидки, руб.")+2}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество показов")+2}))*1000' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Частота")+1, column='Охват технический', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество показов")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Частота")+1}))' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("VTR,%")+1, column='Количество просмотров', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество показов")+1}))*INDIRECT(ADDRESS({i},{filter.columns.get_loc("VTR,%")+1}))' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("CTR,%")+1, column='Количество кликов', value=[
                        f'=IF(INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="1000 показов",INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("CTR,%")+1})),IF(INDIRECT(ADDRESS({i},{filter.columns.get_loc("Единица покупки")+1}))="CPC",INDIRECT(ADDRESS({i},{filter.columns.get_loc("Общее количество единиц")+1})),""))' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Количество кликов")+1, column='CPM, руб.', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Стоимость размещения после скидки, руб.")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество показов")+1}))*1000' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("CPM, руб.")+1, column='CPT, руб.', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Стоимость размещения после скидки, руб.")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Охват технический")+1}))*1000' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("CPT, руб.")+1, column='Стоимость за просмотр', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Стоимость размещения после скидки, руб.")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество просмотров")+1}))*1000' 
                        for i in range(2, len_col+2)])
                    filter.insert(loc=filter.columns.get_loc("Стоимость за просмотр")+1, column='Стоимость за клик, руб.', value=[
                        f'=INDIRECT(ADDRESS({i},{filter.columns.get_loc("Стоимость размещения после скидки, руб.")+1}))/INDIRECT(ADDRESS({i},{filter.columns.get_loc("Количество кликов")+1}))*1000' 
                        for i in range(2, len_col+2)])
                    
                datet = brief.duploaded_at.strftime('%x').replace('/', '.')
                if not os.path.exists(os.path.join(path_plan, f"media/clients/{translit(request.user.username)}")):
                        os.mkdir(os.path.join(path_plan, f"media/clients/{translit(request.user.username)}"))
                if not os.path.exists(os.path.join(path_plan, f"media/clients/{translit(request.user.username)}/{translit(client)}")):
                    os.mkdir(os.path.join(path_plan, f"media/clients/{translit(request.user.username)}/{translit(client)}"))
                filter.to_excel(os.path.join(path_plan, 'media', 'clients', translit(request.user.username), translit(client), f'MP_{datet}.xlsx'), index=None)
                
                '''CORRECTION DMP'''
                wb=load_workbook(os.path.join(path_plan, 'media', 'clients', translit(request.user.username), translit(client), f'MP_{datet}.xlsx'))
                sheet = wb.active
                ''' СЕЗОННИКИ
                for row in sheet[f'H3':f'H{len(season)+1}']:
                    for cell in row:
                        if cell == 'проверить':
                            cell.fill = PatternFill(start_color='ff3333', end_color='ff3333', fill_type='solid')
                '''
                sheet.merge_cells('Y1:Z1')
                sheet.column_dimensions['A'].width = 16
                sheet.column_dimensions['B'].width = 16
                sheet.column_dimensions['C'].width = 16
                sheet.column_dimensions['D'].width = 16
                sheet.column_dimensions['E'].width = 40
                sheet.column_dimensions['F'].width = 20
                sheet.column_dimensions['G'].width = 16
                sheet.column_dimensions['H'].width = 16
                sheet.column_dimensions['I'].width = 16
                sheet.column_dimensions['J'].width = 16
                sheet.column_dimensions['K'].width = 35
                sheet.column_dimensions['L'].width = 35
                sheet.column_dimensions['M'].width = 20
                sheet.column_dimensions['N'].width = 20
                sheet.column_dimensions['O'].width = 16
                sheet.column_dimensions['P'].width = 16
                sheet.column_dimensions['Q'].width = 16
                sheet.column_dimensions['R'].width = 18
                sheet.column_dimensions['S'].width = 14
                sheet.column_dimensions['T'].width = 14
                sheet.column_dimensions['U'].width = 14
                sheet.column_dimensions['V'].width = 14
                sheet.column_dimensions['W'].width = 35
                sheet.column_dimensions['X'].width = 12
                sheet.column_dimensions['Y'].width = 12
                sheet.column_dimensions['Z'].width = 12
                sheet.column_dimensions['AA'].width = 12
                sheet.column_dimensions['AB'].width = 12
                sheet.column_dimensions['AC'].width = 12
                sheet.column_dimensions['AD'].width = 17
                sheet.column_dimensions['AE'].width = 13
                sheet.column_dimensions['AF'].width = 13
                sheet.column_dimensions['AG'].width = 12
                sheet.column_dimensions['AH'].width = 16
                sheet.column_dimensions['AI'].width = 14
                sheet.column_dimensions['AJ'].width = 15
                sheet.column_dimensions['AK'].width = 13
                sheet.column_dimensions['AL'].width = 13
                sheet.column_dimensions['AM'].width = 12
                sheet.column_dimensions['AN'].width = 12
                sheet.column_dimensions['AO'].width = 12
                sheet.column_dimensions['AP'].width = 12
                sheet.column_dimensions['AQ'].width = 12
                sheet.column_dimensions['AR'].width = 16
                sheet.column_dimensions['AS'].width = 16
                sheet.column_dimensions['AT'].width = 16
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True,
                                                   horizontal='center',
                                                   vertical="center")
                        cell.border = Border(top = Side(border_style='thin', color='FF000000'),
                            right = Side(border_style='thin', color='FF000000'),
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))
                font = Font(color="FFFFFFFF")
                HeaderFill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                for cell in list(sheet.iter_rows())[0]:
                    cell.fill = HeaderFill
                    cell.font = font
                for row in list(sheet[f'AK2':f'AK{len_col+2}']+sheet[f'AM2':f'AM{len_col+2}']):
                    for cell in row:
                        #cell.number_format = '"$"#,##0.00_);("$"#,##0.00)' # с рублями
                        pass

                #COMMENTS
                for row in range(2, len_col+2):
                    sheet[f"AD{row}"].comment = Comment(comments_margin[row-2], '.')

                wb.save(os.path.join(path_plan, 'media', 'clients', translit(request.user.username), translit(client), f'MP_{datet}.xlsx'))


                '''
                datet = brief.duploaded_at.strftime('%x').replace('/', '.')
                # In the down def to create a file DMP.xlsx
                path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                #n = 


                hol = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                p = pd.read_excel(os.path.join(hol, n), engine='openpyxl',
                                     header=5)
                e = openpyxl.load_workbook(filename=os.path.join(hol, n), data_only=True)
                ws = e.worksheets[0]
                sheet = e.active

                head_dmp = {} #назв столбцов, где есть примечания и rows
                all_columns_com = {}

                with open(os.path.join(hol, "media/pattern/comments.txt"), 'r') as f:
                    for name in f.read().split():
                        col = column_index_from_string(coordinate_from_string(name)[0])
                        if head_dmp.get(sheet.cell(row=6, column=col).value):
                            head_dmp[sheet.cell(row=6, column=col).value].append(coordinate_from_string(name)[1] - 6)
                        else:
                            head_dmp[sheet.cell(row=6, column=col).value] = [coordinate_from_string(name)[1] - 6]
                        all_columns_com[sheet.cell(row=6, column=col).value] = col


                a = p["Категория Клиента"].tolist()
                b = p["KPI"].tolist()
                video = p['Размер (в пикселях) / Формат'].tolist()
                season = p["Сезонники"].tolist()
                baing = p["Баинговые приоритеты"].tolist()

                u = p["коэф. скидки от 1 (min стоимость плана) до  3 (max стоимость плана) "].tolist()
                k = []
                for i in range(1, len(a)+1):
                    if (a[i-1] == 'Все' or a[i-1] == str(type_act)) and (str(KPI) == '') and (u[i-1]=='1-3' or u[i-1] == discount):
                        k.append(6+i)
                    elif (a[i-1] == 'Все' or a[i-1] == str(type_act)) and (str(KPI) in str(b[i-1])) and (u[i-1]=='1-3' or u[i-1] == discount):
                        k.append(6+i)
                '''
                '''
                baing_d = dict()
                k2 = []

                for i in k:
                    try:
                        if Profile.objects.filter(bying_username=request.user.username) != None:
                            if int(Dmp_priority.objects.get(agency=Profile.objects.filter(bying_username=request.user.username).agency)) not in baing_d:
                                baing_d[int(baing[i-7])]=[i]
                            else:
                                baing_d[int(baing[i-7])].append(i)
                        else:
                            if int(Dmp_priority.objects.get(agency=Profile.objects.filter(manager_username=request.user.username).agency)) not in baing_d:
                                baing_d[int(baing[i-7])]=[i]
                            else:
                                baing_d[int(baing[i-7])].append(i)
                    except ValueError:
                        k2.append(i)

                k=[]

                for i in sorted(baing_d):
                    k.extend(baing_d[i])
                k.extend(k2)
                '''
                '''
                data = dict.fromkeys([i for i in p.columns.ravel()])

                period1 = list(period_c.split('-'))
                period2 = list(period_p.split('-'))

                mon = {1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
                         5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
                         9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"}
                month = []
                if period1[0]<period2[0] and period1[1]==period2[1]:
                    m = 1
                    for i in data:
                        m+=1
                        line = []
                        for j in k:
                            line.append(e['Лист1'].cell(row=j, column=m).value)
                        data[i] = line
                    s = pd.DataFrame(data)

                    if not os.path.exists(os.path.join(hol, f"media/clients/{translit(username)}")):
                        os.mkdir(os.path.join(hol, f"media/clients/{translit(username)}"))
                    if not os.path.exists(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}")):
                        os.mkdir(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}"))
                    s.to_excel(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/DMP_{translit(client)}_{datet}.xlsx"), startrow=1, index=False)
                else:
                    if int(period2[1])<int(period1[1]):
                        period2[1] = int(period2[1])+12
                    for i in range(int(period1[1]), int(period2[1])+1):
                        if i>12:
                            i-=12
                        month.append(mon[int(i)])


                    d = []
                    for j in k:
                        line = []
                        if season[j-7]=='проверить' or season[j-7]=='нет':
                            for i in range(1, len(p.columns.ravel())+1):
                                line.append(e['Лист1'].cell(row=j, column=i).value)
                            d.append(line)
                        else:
                            try:
                                line1 = []
                                for n in list(season[j-7].split()):
                                    if n in month:
                                        line1.append(n)
                                if line1!=[]:
                                    for v in range(1, len(p.columns.ravel())+1):
                                        line.append(e['Лист1'].cell(row=j, column=v).value)
                                    line[7] = line1
                                    d.append(line)
                            except ValueError:
                                for i in range(1, len(p.columns.ravel())+1):
                                    line.append(e['Лист1'].cell(row=j, column=i).value)
                                d.append(line)
                            except AttributeError:
                                for i in range(1, len(p.columns.ravel())+1):
                                    line.append(e['Лист1'].cell(row=j, column=i).value)
                                d.append(line)

                    if len(d)==0:
                        dataclass['er'] = 'нет данных по пункту - Задача, kpi'
                        return render(request, self.template_name, dataclass)
                    for i in range(len(d[0])):
                        m = []
                        for j in range(len(d)):
                            m.append(d[j][i])
                        data[(p.columns.ravel())[i]] = m
                    s = pd.DataFrame(data)


                    if not os.path.exists(os.path.join(hol, f"media/clients/{translit(username)}")):
                        os.mkdir(os.path.join(hol, f"media/clients/{translit(username)}"))
                    if not os.path.exists(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}")):
                        os.mkdir(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}"))

                    s.to_excel(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/DMP_{translit(client)}_{datet}.xlsx"), startrow=1, index=False)
                '''
                ''' This is create brief file for clients'''
                '''

                for i in Brief_pattern.objects.all():
                    n = f'media/{i.file.name}'
                #dataclass['t'] = [hol, n, str(os.path.join(hol, str(n)))]
                #return render(request, self.template_name, dataclass)
                wb = openpyxl.load_workbook(filename=os.path.join(hol, n), data_only=True)
                ws = wb.worksheets[0]
                sheet = wb.active
                bd = Brief.objects.filter(username=username, client=client,
                                       product=product, name_rk=name_rk)[::-1][0]
                sheet['C3'] = bd.client
                sheet['C4'] = bd.product
                sheet['C5'] = bd.name_rk
                sheet['C6'] = bd.posad
                sheet['C7'] = bd.description
                sheet['C8'] = bd.competitors
                sheet['C9'] = bd.type_act
                sheet['C11'] = bd.country
                sheet['C12'] = bd.region
                sheet['C14'] = bd.gender
                sheet['C15'] = bd.age
                sheet['C16'] = bd.interes
                sheet['C17'] = bd.income
                sheet['C18'] = bd.rek
                sheet['C20'] = bd.materials
                sheet['C21'] = str(bd.period_c) + " - " + str(bd.period_p)
                sheet['C22'] = bd.budget
                sheet['C23'] = bd.KPI
                sheet['C24'] = bd.plan
                sheet['C25'] = bd.who_prep_materials
                sheet['H3'] = bd.discount
                sheet['H4'] = bd.AK
                sheet['H5'] = bd.DCM
                try:
                    im = openpyxl.drawing.image.Image(bd.img)
                    im.height = 250
                    im.width = 250
                    ws.add_image(im, 'H6')
                except:
                    pass

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True,vertical='top')
                wb.save(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/brief_{translit(client)}_{datet}.xlsx"))
                path2 = join('clients', translit(username), translit(client), f"brief_{translit(client)}_{datet}.xlsx")
                '''

                '''This is correction DMP'''
                '''
                wb=load_workbook(os.path.join(hol, f"media/clients/{username}/{client}/DMP_{name_rk}.xlsx"))
                sheet = wb.active

                p = pd.read_excel(os.path.join(hol, f"media/clients/{username}/{client}/DMP_{name_rk}.xlsx"),
                                     header=1)
                season = p['Сезонники'].tolist()
                for row in sheet[f'H3':f'H{len(season)+1}']:
                    for cell in row:
                        if cell == 'проверить':
                            cell.fill = PatternFill(start_color='ff3333', end_color='ff3333', fill_type='solid')
                sheet.column_dimensions['B'].width = 10
                sheet.column_dimensions['C'].width = 8
                sheet.column_dimensions['D'].width = 17
                sheet.column_dimensions['E'].width = 11
                sheet.column_dimensions['F'].width = 145
                sheet.column_dimensions['G'].width = 11
                sheet.column_dimensions['H'].width = 14
                sheet.column_dimensions['I'].width = 65
                sheet.column_dimensions['J'].width = 13
                sheet.column_dimensions['K'].width = 13
                sheet.column_dimensions['L'].width = 13
                sheet.column_dimensions['M'].width = 15
                sheet.column_dimensions['N'].width = 15
                sheet.column_dimensions['O'].width = 19
                sheet.column_dimensions['P'].width = 45
                sheet.column_dimensions['Q'].width = 15
                sheet.column_dimensions['R'].width = 15
                sheet.column_dimensions['S'].width = 25
                sheet.column_dimensions['T'].width = 15
                sheet.column_dimensions['U'].width = 25
                sheet.column_dimensions['V'].width = 19
                sheet.column_dimensions['W'].width = 35
                sheet.column_dimensions['X'].width = 12
                sheet.column_dimensions['Y'].width = 12
                sheet.column_dimensions['Z'].width = 12
                sheet.column_dimensions['AA'].width = 12
                sheet.column_dimensions['AB'].width = 12
                sheet.column_dimensions['AC'].width = 8
                sheet.column_dimensions['AD'].width = 17
                sheet.column_dimensions['AE'].width = 11
                sheet.column_dimensions['AF'].width = 9
                sheet.column_dimensions['AG'].width = 12
                sheet.column_dimensions['AH'].width = 14
                sheet.column_dimensions['AI'].width = 14
                sheet.column_dimensions['AJ'].width = 15
                sheet.column_dimensions['AK'].width = 13
                sheet.column_dimensions['AL'].width = 13
                sheet.column_dimensions['AM'].width = 12
                sheet.column_dimensions['AN'].width = 12
                sheet.column_dimensions['AO'].width = 12
                sheet.column_dimensions['AP'].width = 12
                sheet.column_dimensions['AQ'].width = 12
                sheet.column_dimensions['AR'].width = 12
                sheet.column_dimensions['AS'].width = 12
                sheet.column_dimensions['AT'].width = 12
                sheet.column_dimensions['AU'].width = 12
                sheet.column_dimensions['AV'].width = 12
                sheet.column_dimensions['AW'].width = 12
                sheet.column_dimensions['AX'].width = 12
                sheet.column_dimensions['AY'].width = 20
                sheet.column_dimensions['AZ'].width = 12
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True,vertical='top',
                                                   horizontal='center')
                        cell.border = Border(top = Side(border_style='thin', color='FF000000'),
                            right = Side(border_style='thin', color='FF000000'),
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))
                wb.save(os.path.join(hol, f"media/clients/{username}/{client}/DMP_{name_rk}.xlsx"))
                '''

                '''This is create mp'''
                '''
                p = pd.read_excel(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/DMP_{translit(client)}_{datet}.xlsx"), engine='openpyxl',
                                  header=None, skiprows=2, usecols = [1, 2, 3, 4, 5, 6,
                                                                    8, 9, 11, 12, 13,
                                                                    14, 15, 16, 17,
                                                                    18, 19, 20, 21, 22,
                                                                    23, 24, 25, 26, 27, 29, 30, 31])
                frequency = pd.read_excel(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/DMP_{translit(client)}_{datet}.xlsx"), engine='openpyxl',
                                          header=None, skiprows=2, usecols = [37, 39, 41])

                try:
                    f = Profile.objects.get(bying_username=username)
                except ObjectDoesNotExist:
                    f = Profile.objects.get(manager_username=username)
                report_link = f.report_common

                report = pd.read_excel(os.path.join(hol, f"{report_link}"), engine='openpyxl',
                                  header=None, skiprows=6)


                fr = frequency.to_dict(orient='list')
                b = p.to_dict(orient='list')
                report = report.to_dict(orient='list')

                height = len(b[4])
                h1 = height
                '''
                '''лиды'''
                '''
                lids = [''] * height
                try:
                    for i in range(0, len(b[21])):
                        if client in report[1][::-1]:
                            step = len(report[1])+1
                            for j in range(len(report[1])):
                                if client in report[1][step-1::-1]:
                                    count_find = height - list(report[1])[step-1::-1].index(client)
                                    if report[3][count_find] == b[21][i]: #Site
                                        lids[i]=report[7][count_find]
                                        break
                                    else:
                                        step = list(report[1])[step-1::-1].index(client) #indect the last enterring client
                                else:
                                    break
                except:
                    pass
                '''

                '''ctr and vtr'''
                '''
                ctr = [''] * height
                vtr = [''] * height
                try:
                    for i in range(0, len(b[21])):
                        if client in report[1][::-1]:
                            step = len(report[1])+1
                            for j in range(len(report[1])):
                                if client in report[1][step-1::-1]:
                                    count_find = height - list(report[1])[step-1::-1].index(client)
                                    if report[3][count_find] == b[21][i]: #Site and Format
                                        if b[23][i] in report[4][count_find]:
                                            ctr[i]=report[6][count_find]
                                            vtr[i]=report[5][count_find]
                                        else:
                                            step = list(report[1])[step-1::-1].index(client) #indect the last enterring client
                                    else:
                                        step = list(report[1])[step-1::-1].index(client)
                                else:
                                    break
                        elif b[1][i] in report[2][::-1]:
                            step = len(report[1])+1
                            for j in range(len(report[1])):
                                if b[1][i] in report[2][step-1::-1]:
                                    count_find = height - list(report[2])[step-1::-1].index(b[1][i]) #категория клиента
                                    if report[3][count_find]==b[21][i]: # Site
                                        if b[23][i] in report[4][count_find]:
                                            ctr[i]=float(report[6][count_find])*0.95
                                            vtr[i]=float(report[5][count_find])*0.95
                                        else:
                                            step = list(report[2])[step-1::-1].index(b[1][i])
                                    else:
                                        step = list(report[2])[step-1::-1].index(b[1][i])
                                else:
                                    break
                        elif b[21][i] in report[3][::-1]: #Site
                            step = len(report[1])+1
                            for j in range(len(report[1])):
                                if b[21][i] in report[3][step-1::-1]:
                                    count_find = height - list(report[3])[step-1::-1].index(b[21][i])
                                    # Format
                                    if b[23][i] in report[4][count_find]:
                                        ctr[i]=float(report[6][count_find])*0.92
                                        vtr[i]=float(report[5][count_find])*0.92
                                    else:
                                        step = list(report[3])[step-1::-1].index(b[21][i])
                                else:
                                    break
                except:
                    pass
                if '' in ctr:
                    for i in range(0, len(b[21])):
                        if ctr[i] == '':
                            for w in range(height):
                                if b[21][w]==b[21][i] and b[23][w]==b[23][i] and i!=w:
                                    ctr[i]=fr[41][w]*90/100
                                    vtr[i]=fr[39][w]*90/100
                                    break





                b[35] = b.pop(11)

                b[20] = [i for i in range(1, height+1)]
                b[33] = b.pop(30)
                b[34] = b.pop(31)
                b[32] = b.pop(29)
                b[30] = b.pop(27)
                b[29] = [f'=COUNT(AV{i}:DC{i})' for i in range(13, height+13)]
                b[31] = [f'=AB{i}/Y{i}' for i in range(13, height+13)]
                b[27] = b.pop(25)
                b[26] = b.pop(24)
                b[24] = b.pop(23)
                b[23] = b.pop(22)
                b[22] = b.pop(21)
                b[21] = [f'=S{i}' for i in range(13, height+13)]

                b[25] = ['']*height #длит видео

                b = dict(sorted(b.items(), key=lambda x: x[0]))
                b[37] = [f'=IF(OR(X{i}="1000 показов",X{i}="клики",X{i}="engagement",X{i}="вовлечение",X{i}="просмотры"),IF(X{i}="клики",AG{i}*1000/AI{i},IF(OR(X{i}="engagement",X{i}="просмотры",X{i}="вовлечение"),AG{i}*1000/AI{i},AC{i}*AD{i}*(1-AE{i}))),IF(ISERR(AC{i}*AD{i}/AI{i}*1000*(1-AE{i})),0,AC{i}*AD{i}*AB{i}*(1-AE{i})/AI{i}*1000))' for i in range(13, height+13)]
                b[38] = [f'=IF(X{i}="клики",AC{i}*AD{i}*(1-AE{i})*AO{i},IF(OR(X{i}="просмотры",X{i}="engagement",X{i}="вовлечение"),AB{i}*AC{i}*AD{i}*(1-AE{i}),IF(OR(X{i}="пакет",X{i}="неделя",X{i}="день",X{i}="месяц",X{i}="единица",X{i}="единиц"),AC{i}*AD{i}*(1-AE{i})*AB{i},AB{i}*AF{i})))' for i in range(13, height+13)]
                b[39] = [f'=AG{i}*1.2' for i in range(13, height+13)]
                b[40] = [f'=AB{i}*1000' for i in range(13, height+13)]
                b[41] = fr[37]
                b[42] = [f'=AI{i}/AJ{i}' for i in range(13, height+13)]
                b[43] = ['']*height
                b[44] = [f'=AB{i}' for i in range(13, height+13)]
                b[45] = ctr
                b[46] = [f'=AI{i}*AN{i}' for i in range(13, height+13)]
                b[47] = [f'=AG{i}/AI{i}*1000' for i in range(13, height+13)]
                b[48] = [f'=AG{i}/AK{i}*1000' for i in range(13, height+13)]
                b[49] = [f'=AG{i}/AM{i}' for i in range(13, height+13)]
                b[50] = [f'=AG{i}/AO{i}' for i in range(13, height+13)]
                b[51] = lids
                b[52] = [f'=AG{i}/AT{i}' for i in range(13, height+13)]

                u=pd.DataFrame(b)

                for i in Media_plan.objects.all():
                    media_plan = i.file.name
                wb = openpyxl.load_workbook(filename=os.path.join(hol, f"media/{media_plan}"))
                w = wb.worksheets[0]
                sheet = wb.active

                for r in dataframe_to_rows(u, index=None, header=None):
                    w.append(r)




                g = []
                '''
                '''
                for i in [bd.duration1, bd.duration2, bd.duration3]:
                    if i!='':
                        g.append(i)
                if len(g)==0:
                    for r in dataframe_to_rows(u, index=None, header=None):
                        w.append(r)

                else:
                    for during in g:
                        b[25] = [during]*height      # the lasting of video(use in the format)

                        vtr = [''] * height
                        try:
                            for i in range(0, len(b[21])):
                                if client in report[1][::-1]:
                                    step = len(report[1])+1
                                    for j in range(len(report[1])):
                                        if client in report[1][step-1::-1]:
                                            count_find = height - list(report[1])[step-1::-1].index(client)
                                            if report[3][count_find] == b[21][i] and during in report[4][count_find]: #Site and lasting
                                                vtr[i]=report[5][count_find]
                                                break
                                            else:
                                                step = list(report[1])[step-1::-1].index(client)
                                        else:
                                            break
                                if client in report[1][::-1] and vtr[i] == '':
                                    step = len(report[1])+1
                                    for j in range(len(report[1])):
                                        if client in report[1][step-1::-1]:
                                            count_find = height - list(report[1])[step-1::-1].index(client)
                                            if during in report[4][count_find]: #lasting
                                                try:
                                                    vtr[i]=float(report[5][count_find]*0.9)
                                                except TypeError:
                                                    pass
                                            else:
                                                step = list(report[1])[step-1::-1].index(client)
                                        else:
                                            break
                                elif b[1][i] in report[2][::-1] and vtr[i] == '': #категория клиента
                                    step = len(report[1])+1
                                    for j in range(len(report[1])):
                                        if b[1][i] in report[2][step-1::-1]:
                                            count_find = height - list(report[2])[step-1::-1].index(b[1][i]) #категория клиента
                                            if report[3][count_find]==b[21][i]: # Site
                                                try:
                                                    vtr[i]=float(report[5][count_find]*0.88)
                                                except TypeError:
                                                    pass
                                            else:
                                                step = list(report[2])[step-1::-1].index(b[1][i])
                                        else:
                                            break
                                if b[21][i] in report[3][::-1] and vtr[i] == '': #Site
                                    step = len(report[1])+1
                                    for j in range(len(report[1])):
                                        if b[21][i] in report[3][step-1::-1]:
                                            count_find = height - list(report[3])[step-1::-1].index(b[21][i])
                                            if report[3][count_find]==b[21][i] and during in report[4][count_find]:
                                                try:
                                                    vtr[i]=float(report[5][count_find]*0.85)
                                                except TypeError:
                                                    pass
                                            else:
                                                step = list(report[2])[step-1::-1].index(b[1][i])
                                        else:
                                            break
                        except:
                            pass
                        if '' in vtr:
                            for i in range(0, len(b[21])):
                                if vtr[i] == '':
                                    for j in range(height):
                                        if b[1][j]==b[1][i] and b[21][j]==b[21][i] and during==b[23][i] and i!=j:
                                            vtr[i]=fr[39][j]*90/100
                                            break


                        b[43] = vtr
                        u=pd.DataFrame(b)
                        for r in dataframe_to_rows(u, index=None, header=None):
                            w.append(r)
                        '''
                '''
                formula = '1000 показов, клики, пакет, просмотры, engagement, вовлечение, неделя, месяц, единица, единиц, день'
                dv = DataValidation(type='list', formula1='"{}"'.format(formula), allow_blank=True)
                sheet.add_data_validation(dv)
                dv.add(f'X13:X{height+12}')

                formula1 = 'день, дней, дня, неделя, недели, недель, месяц, месяца, месяцев, единица, единиц, единицы'
                dv = DataValidation(type='list', formula1='"{}"'.format(formula1), allow_blank=True)
                sheet.add_data_validation(dv)
                dv.add(f'Z13:Z{height+12}')

                sheet['T2'] = bd.client
                sheet['T3'] = bd.product
                sheet['T4'] = bd.posad
                sheet['T5'] = str(bd.gender) + ", " + str(bd.age) + ", " + str(bd.interes)
                sheet['T6'] = str(bd.country) + ", " + str(bd.region)
                sheet['T7'] = bd.KPI

                for row in list(sheet)[12:]:
                    for cell in row:
                        cell.border = Border(top = Side(border_style='thin', color='FF000000'),
                            right = Side(border_style='thin', color='FF000000'),
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))

                sheet[f'AE{height+13}'] = 'Итого:'
                sheet[f'AF{height+13}'] = f'=SUMIF(AI13:AI{height+12},">0",AG13:AG{height+12})/AI{height+13}*1000'
                sheet[f'AG{height+13}'] = f'=SUM(AG13:AG{height+12})'
                sheet[f'AH{height+13}'] = f'=SUM(AH13:AH{height+12})'
                sheet[f'AI{height+13}'] = f'=SUM(AI13:AI{height+12})'
                sheet[f'AJ{height+13}'] = f'=SUMIF(AK13:AK{height+12},">0",AI13:AI{height+12})/AK{height+13}'
                sheet[f'AK{height+13}'] = f'=SUM(AK13:AK{height+12})*0.8'
                sheet[f'AL{height+13}'] = f'=SUMIF(AI13:AI{height+12},">0",AM13:AM{height+12})/AI{height+13}'
                sheet[f'AM{height+13}'] = f'=SUM(AM13:AM{height+12})'
                sheet[f'AN{height+13}'] = f'=SUMIF(AI13:AI{height+12},">0",AO13:AO{height+12})/AI{height+13}'
                sheet[f'AO{height+13}'] = f'=SUM(AO13:AO{height+12})'
                sheet[f'AP{height+13}'] = f'=SUMIF(AI13:AI{height+12},">0",AG13:AG{height+12})/AI{height+13}*1000'
                sheet[f'AQ{height+13}'] = f'=SUMIF(AK13:AK{height+12},">0",AG13:AG{height+12})/AK{height+13}*1000'
                sheet[f'AR{height+13}'] = f'=SUMIF(AM13:AM{height+12},">0",AG13:AG{height+12})/AM{height+13}'
                sheet[f'AS{height+13}'] = f'=SUMIF(AO13:AO{height+12},">0",AG13:AG{height+12})/AO{height+13}'
                sheet[f'AT{height+13}'] = f'=SUM(AT13:AM{height+12})'
                sheet[f'AU{height+13}'] = f'=SUMIF(AT13:AT{height+12},">0",AG13:AG{height+12})/AT{height+13}'
                sheet[f'AV{height+13}'] = f'=SUMIF(AU13:AU{height+12},">0",AG13:AG{height+12})/AU{height+13}'
                sheet[f'AC{height+14}'] = 'Tracker'
                sheet[f'AC{height+15}'] = 'Итого медиа бюджет'
                sheet[f'AC{height+16}'] = 'АК'
                sheet[f'AC{height+17}'] = 'НДС'
                sheet[f'AC{height+18}'] = 'Производство ролика, с НДС'
                sheet[f'AC{height+19}'] = 'Итого (с учётом НДС и АК)'
                sheet[f'AF{height+16}'] = bd.AK
                sheet[f'AF{height+17}'] = '20%'

                sheet[f'AG{height+14}'] = f'=(AI{height+13}*2.5)*1.5/1000'
                sheet[f'AG{height+15}'] = f'=SUM(AG{height+13}:AG{height+14})'
                sheet[f'AG{height+16}'] = f'=AG{height+15}*AF{height+16}'
                sheet[f'AG{height+17}'] = f'=((AG{height+15})+AG{height+16})*AF{height+17}+AC{height+14}'
                sheet[f'AG{height+18}'] = '0.00р'
                sheet[f'AG{height+19}'] = f'=SUM(AG{height+15}:AG{height+18})'

                font = Font(color="FFFFFFFF")
                HeaderFill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                for row in sheet[f'Q{height+13}':f'AV{height+13}']:
                    for cell in row:
                        cell.fill = HeaderFill
                        cell.font = font
                for row in sheet[f'AC{height+19}':f'AF{height+19}']:
                    for cell in row:
                        cell.border = Border(
                            bottom = Side(border_style='thin', color='FF000000'))
                for row in sheet[f'AC{height+14}':f'AC{height+19}']:
                    for cell in row:
                        cell.border = Border(
                            left = Side(border_style='thin', color='FF000000'))

                sheet[f'AC{height+19}'].border = Border(top = Side(border_style='thin', color='FF000000'),
                    bottom = Side(border_style='thin', color='FF000000'),
                    left = Side(border_style='thin', color='FF000000'))

                for row in sheet[f'AG{height+14}':f'AG{height+19}']:
                    for cell in row:
                        cell.border = Border(top = Side(border_style='thin', color='FF000000'),
                            right = Side(border_style='thin', color='FF000000'),
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))
                for row in list(sheet[f'AG13':f'AG{height+19}']+sheet[f'AC13':f'AC{height+13}']
                                +sheet[f'AH13':f'AH{height+13}']+sheet[f'AP13':f'AP{height+13}']
                                +sheet[f'AQ13':f'AQ{height+13}']+sheet[f'AR13':f'AR{height+13}']
                                +sheet[f'AS13':f'AS{height+13}']+sheet[f'AU13':f'AU{height+13}']
                                +sheet[f'AF13':f'AF{height+13}']):
                    for cell in row:
                        cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'
                for row in list(sheet[f'AA13':f'AA{height+19}']+sheet[f'AB13':f'AB{height+13}']
                                +sheet[f'AI13':f'AI{height+13}']+sheet[f'AK13':f'AK{height+13}']
                                +sheet[f'AM13':f'AM{height+13}']+sheet[f'AO13':f'AO{height+13}']):
                    for cell in row:
                        cell.number_format = '#,##0_);(#,##0)'
                for row in list(sheet[f'AL13':f'AL{height+19}']+sheet[f'AN13':f'AN{height+13}']
                                +sheet[f'AE13':f'AE{height+13}']):
                    for cell in row:
                        cell.number_format = '0.00%'
                for row in list(sheet.iter_rows())[12:]:
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True,vertical='top')
                for i in range(13, height+13):
                    sheet.row_dimensions[i].height = 70
                '''
                ''' Сезонники и тайминг '''
                '''
                p = pd.read_excel(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/DMP_{translit(client)}_{datet}.xlsx"), engine='openpyxl',
                                     header=1)
                season = p['Сезонники'].tolist()
                season2 = {}
                for i in range(48, 104):
                    sheet.column_dimensions[get_column_letter(i)].width = 0.01
                    if sheet.cell(row=10, column=i).value!=None:
                        season2[sheet.cell(row=10, column=i).value] = i

                for i in range(104, 108):
                    sheet.column_dimensions[get_column_letter(i)].width = 0.01

                #sheet.column_dimensions['V'].width = 0.01

                for i in range(13, len(season)+13):
                    if season[i-13]=='проверить' or season[i-13]=='нет':
                        for s in range(48, 108):
                            sheet.cell(row=i, column=s).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                    else:
                    for h in list(season[i-13][1:-1].replace("'", "").split(', ')):
                        f = season2[h]
                        for k in range(5):
                            sheet.cell(row=i, column=f+k).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                    """если год"""
                    if period1[0]<period2[0] and period1[1]==period2[1]:
                        for s in range(48, 108):
                            sheet.cell(row=i, column=s).value = 1
                            sheet.cell(row=i, column=s).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            """если 2 месяца"""
                    elif int(period2[1])-int(period1[1])==1:
                        for a in range(math.ceil(int(period1[2])/7)-1, 5):
                            sheet.cell(row=i, column=season2[month[0]]+a).value = 1
                            sheet.cell(row=i, column=season2[month[0]]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            sheet.column_dimensions[get_column_letter(season2[month[0]]+a)].width = 8
                        for a in range(math.ceil(int(period2[2])/7)):
                            sheet.cell(row=i, column=season2[month[1]]+a).value = 1
                            sheet.cell(row=i, column=season2[month[1]]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            sheet.column_dimensions[get_column_letter(season2[month[1]]+a)].width = 8
                            """если больше 2 месяцев"""
                    elif int(period2[1])-int(period1[1])>1:
                        for a in range(math.ceil(int(period1[2])/7)-1, 5):
                            sheet.cell(row=i, column=season2[month[0]]+a).value = 1
                            sheet.cell(row=i, column=season2[month[0]]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            sheet.column_dimensions[get_column_letter(season2[month[0]]+a)].width = 8
                        for g in month[1:-1]:
                            for a in range(5):
                                sheet.cell(row=i, column=season2[g]+a).value = 1
                                sheet.cell(row=i, column=season2[g]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                                sheet.column_dimensions[get_column_letter(season2[g]+a)].width = 8
                        for a in range(math.ceil(int(period2[2])/7)):
                            sheet.cell(row=i, column=season2[month[-1]]+a).value = 1
                            sheet.cell(row=i, column=season2[month[-1]]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            sheet.column_dimensions[get_column_letter(season2[month[-1]]+a)].width = 8
                            """если 1 месяц"""
                    else:
                        for a in range(math.ceil(int(period1[2])/7)-1, math.ceil(int(period2[2])/7)):
                            sheet.cell(row=i, column=season2[month[0]]+a).value = 1
                            sheet.cell(row=i, column=season2[month[0]]+a).fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
                            sheet.column_dimensions[get_column_letter(season2[month[0]]+a)].width = 8


                try:
                    for j in range(13, height+13):
                        for k in range(len(report[6])-1, 0, -1):
                            if report[1][k]==client and report[6][k]==b[21][j-10] and b[24][j-10]==report[9][k]:
                                sheet.cell(row=j, column=115).value = report[59][k]
                                sheet.cell(row=j, column=116).value=report[60][k]
                                sheet.cell(row=j, column=117).value=report[61][k]
                                sheet.cell(row=j, column=118).value=report[62][k]
                                break
                except:
                    pass
                '''
                ''' comments '''
                '''
                for i in Dmp.objects.all():
                    n = i.file.url[1:]
                wb2 = openpyxl.load_workbook(filename=os.path.join(hol, n), data_only=True)
                w2 = wb2.worksheets[0]
                sheet2 = wb2.active


                for name, row in head_dmp.items():
                    val_mp_head1 = [i.value for i in list(sheet.iter_rows())[11]]
                    val_mp_head2 = [i.value for i in list(sheet.iter_rows())[10]]
                    val_mp_head3 = [i.value for i in list(sheet.iter_rows())[9]]
                    if name in val_mp_head1:
                        for num in row:
                            sheet.cell(row=num+12, column=val_mp_head1.index(name)+1).comment = Comment(sheet2.cell(row=num+6, column=all_columns_com[name]).comment, '.')
                    elif name in val_mp_head2:
                        for num in row:
                            sheet.cell(row=num+12, column=val_mp_head2.index(name)+1).comment = Comment(sheet2.cell(row=num+6, column=all_columns_com[name]).comment, '.')
                    elif name in val_mp_head3:
                        for num in row:
                            sheet.cell(row=num+12, column=val_mp_head3.index(name)+1).comment = Comment(sheet2.cell(row=num+6, column=all_columns_com[name]).comment, '.')




                wb.save(os.path.join(hol, f"media/clients/{translit(username)}/{translit(client)}/mp_{translit(client)}_{datet}.xlsx"))
                '''
                '''
                wb2 = openpyxl.load_workbook(filename=os.path.join(hol, f"media/pattern/buying.xlsx"))
                w2 = wb2.worksheets[0]
                sheet2 = wb2.active
                f = pd.read_excel(os.path.join(hol, f"media/clients/{username}/{client}/mp_{client}_{datet}.xlsx"), engine='openpyxl',
                                     header=None, skiprows=12, usecols=(18, 23, 27, 28, 29, 30))

                price_b_s = []
                for i in range(h1):
                    if f[23][i]=='1000 показов':
                        try:
                            price = int(f[27][i])*int(f[28][i])/1000
                            if f[29][i]!=None:
                                price*=int(f[29][i])
                            if f[30][i]!=None:
                                price*=(1-int(f[30][i]))
                            price_b_s.append(price)
                        except ValueError:
                            price_b_s.append('')
                    else:
                        try:
                            price = int(f[27][i])*int(f[28][i])
                            if f[29][i]!=None:
                                price*=int(f[29][i])
                            if f[30][i]!=None:
                                price*=(1-int(f[30][i]))
                            price_b_s.append(price)
                        except ValueError:
                            price_b_s.append('')
                f1 = {0: f[18][:h1], 1: f[18][:h1], 2: price_b_s}
                bu = pd.DataFrame(f1)

                bu.to_excel(os.path.join(hol, f"media/pattern/buying.xlsx"), header=None, index=None)

                dataset = Dataset()
                f = pd.read_excel(os.path.join(hol, f"media/pattern/buying.xlsx"), engine='openpyxl',
                                     header=None)
                import_data = dataset.load(pd.DataFrame(f))
                for k in import_data:
                    try:
                        a = Bying.objects.get(sell=k[0],site=k[1])
                        a.phact = float(k[2])+float(a.phact)
                        a.save()
                    except ObjectDoesNotExist:
                        value = Bying(None, k[0], k[1], None, k[2])
                        value.save()
                        '''
                '''
                path = join('clients', translit(username), translit(client), f"DMP_{translit(client)}_{datet}.xlsx")
                path3 = join('clients', translit(username), translit(client), f"mp_{translit(client)}_{datet}.xlsx")

                count = All_file.objects.create(username=username, client=client,
                                      name_rk=name_rk, dmp=path, brief=path2,
                                      mp=path3)
                '''
                All_file.objects.create(user_id=request.user.id, username=username, client=client,
                                      name_rk=name_rk, mp=os.path.join('media', 'clients', translit(request.user.username), translit(client), f'MP_{datet}.xlsx'))
                return redirect('exel:calculate')
            return render(request, self.template_name, dataclass)

def calculate(request):
    if request.user.is_authenticated:
        data = {
           'file': All_file.objects.filter(user_id=request.user.id).order_by('-id')[0]
           }
        return render(request, 'prepare_calculation/calculate.html', data)

class Buying(TemplateView):
    template_name = 'buying.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
            # ADD new info about sellers and sites on dict but without saveing
            users_id = Profile_client.objects.all().values_list('user_id', flat=True)
            users_of_profile = [User.objects.get(id=id) for id in users_id]
            for seller in users_of_profile:
                data_of_sellers_from_table = pd.read_excel(os.path.join(os.path.dirname(path), 'sites', translit(seller.username), 'second_part.xlsx'), engine='openpyxl')
                try:
                    for site in data_of_sellers_from_table['Сайт'].tolist():
                        try:
                            BuyingDB.objects.get(seller=seller.username, site=site)
                        except ObjectDoesNotExist:
                            BuyingDB.objects.create(seller=seller.username, site=site)
                except:
                    continue
            

            
            data = {
                'data': BuyingDB.objects.all(),
                }

            if request.method=='POST' and 'form_update' in request.POST:
                sellers = request.POST.getlist('seller')
                sites = request.POST.getlist('site')
                selles = request.POST.getlist('sell')
                buying_priorities = request.POST.getlist('buying_priority')
                print(selles)
                for i in range(len(sellers)):
                    buying_db = BuyingDB.objects.get(seller=sellers[i], site=sites[i])
                    buying_db.sell = selles[i]
                    buying_db.buying_priority = buying_priorities[i]
                    buying_db.save()

            return render(request, self.template_name, data)
        else:
            return redirect('exel:login')

def report(request):
    if request.user.is_authenticated:
        data = {
            'files': ReportFile.objects.filter(user_id=request.user.id),
            'form': ReportForm(),
           }
        if request.method == 'POST':
            file_form = FileModelForm(request.POST, request.FILES)
            files = request.FILES.getlist('file') #field name in model
            if file_form.is_valid():
                for f in files:
                    ReportFile(file=f, user_id=request.user.id, username=request.user.username).save()
        data['form'] = ReportForm()
        return render(request, 'report.html', data)
    else:
        return redirect('exel:login')



class Download_calc(TemplateView):
    template_name = 'download_calc.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            username = request.user.username
            data = {
                'files': Client.objects.filter(username=username)[::-1],
                'name_client': Brief.objects.filter(username=username)[::-1],
                'list_RK': Brief.objects.filter(username=username)[::-1],
                'form': ClientForm,
                }
            if request.method == 'POST':
                client = request.POST.get('name_client')
                title_rk = request.POST.get('title_rk')
                form = ClientForm(request.POST, request.FILES)
                if form.is_valid():
                    ex = request.FILES.get('calculation')
                    h = Brief.objects.filter(username=username)[::-1][0]
                    h.name_rk=title_rk
                    h.save()
                    h = All_file.objects.filter(username=username)[::-1][0]
                    h.name_rk=title_rk
                    h.save()
                    Client.objects.create(username=username, calculation=ex,
                                              client=client, name_rk=title_rk)
                    data['files'] = Client.objects.filter(username=username)[::-1]
                    return render(request, self.template_name, data)
                else:
                    messages.error(request, f'ERROR: Format of uploaded file: {ex.name} is NOT supported !')
            else:
                form = ClientForm
            data['form'] = form
            return render(request, self.template_name, data)
        else:
            return redirect('exel:login')



def not_cleared(request, pk):
    if request.user.is_authenticated:
        username = request.user.username
        try:
            m = get_object_or_404(All_file, pk=pk)
            data = {
                'client': m.client,
                'mp': m.mp,
                'brief': m.brief,
                'form': CommentForm(initial={
                    'name_rk': m.name_rk,
                    'comments': m.comments,
                        }),
                }
            if request.method == 'POST':
                form = CommentForm(request.POST or None)
                if form.is_valid():
                    ex = request.FILES.get('presentation')
                    mp = request.FILES.get('mp')
                    comment = request.POST.get('comments')
                    rk = request.POST.get('name_rk')
                    if m.name_rk!=rk:
                        Client.objects.filter(username=username,
                                                  name_rk=m.name_rk).update(name_rk=rk)
                        Complete.objects.filter(username=username,
                                                  name_rk=m.name_rk).update(name_rk=rk)
                        Brief.objects.filter(username=username,
                                                  name_rk=m.name_rk).update(name_rk=rk)
                        All_file.objects.filter(username=username,
                                                  name_rk=m.name_rk).update(name_rk=rk)
                        Cleared.objects.filter(username=username,
                                                  name_rk=m.name_rk).update(name_rk=rk)
                        Feed.objects.filter(name_rk=m.name_rk).update(name_rk=rk)
                    if ex==None and mp!=None:
                        try:
                            d = Cleared.objects.get(username=username, name_rk=rk,
                                              client=m.client)
                            m.mp = mp
                            m.comments = comment
                            m.save()
                            d.mp = mp
                            d.save()
                            h = Cleared.objects.filter(username=username,
                                                  name_rk=m.name_rk)
                            h.update(mp=mp)
                            h = All_file.objects.filter(username=username,
                                                      name_rk=m.name_rk)
                            h.update(mp=mp)
                        except:
                            data['error'] = 'Заполните все поля'
                            return render(request, 'but_not_cleared.html', data)

                    elif mp==None and ex==None:
                        try:
                            d = Cleared.objects.get(username=username, name_rk=rk,
                                                  client=m.client)
                            All_file.objects.filter(username=username,
                                                  name_rk=rk).update(comments=comment)
                        except:
                            data['error'] = 'Заполните все поля'
                            return render(request, 'but_not_cleared.html', data)
                    else:
                        All_file.objects.filter(username=username,
                                                  name_rk=rk).update(presentation=ex, comments=comment)
                        b = Brief.objects.filter(name_rk=rk)[::-1][0]
                        try:
                            d = Cleared.objects.get(username=username, name_rk=rk,
                                                  client=m.client)
                            d.mp = mp
                            d.save()
                        except ObjectDoesNotExist:
                            Cleared.objects.create(username=username, name_rk=rk,
                                          client=m.client, mp=m.mp, landing=b.posad)
                    h = Client.objects.filter(username=username, name_rk=rk)
                    h.update(comments=comment)
                return main(request)
            return render(request, 'but_not_cleared.html', data)
        except ObjectDoesNotExist:
            pass
    else:
        return redirect('exel:login')


def cleared(request, pk):
    if request.user.is_authenticated:
        username = request.user.username
        try:
            f = Cleared.objects.get(username=username,
                                                 name_rk=name_rk)
            data = {
               'file': f,
               'report': Report.objects.all(),
               'form2': ReportForm,
               'form1': ClearForm(initial={
                   'name_rk': name_rk,
                   'comments': f.comments,
                   'landing': f.landing,
                   'access': f.access
                       }),
               }
            if request.method=='POST' and 'form1' in request.POST:
                form1 = ClearForm(request.POST or None)
                if form1.is_valid():
                    mp = request.FILES.get('mp')
                    comments = request.POST.get('comments')
                    rk = request.POST.get('name_rk')
                    landing = request.POST.get('landing')
                    access = request.POST.get('access')
                    if name_rk!=rk:
                        Client.objects.filter(username=username,
                                                  name_rk=name_rk).update(name_rk=rk)
                        Complete.objects.filter(username=username,
                                                  name_rk=name_rk).update(name_rk=rk)
                        Brief.objects.filter(username=username,
                                                  name_rk=name_rk).update(name_rk=rk)
                        All_file.objects.filter(username=username,
                                                  name_rk=name_rk).update(name_rk=rk)
                        Feed.objects.filter(name_rk=name_rk).update(name_rk=rk)
                    if mp==None:
                        f.name_rk = rk
                        f.comments = comments
                        f.landing = landing
                        f.access = access
                        f.save()
                    else:
                        f.name_rk = rk
                        f.mp = mp
                        f.comments = comments
                        f.save()
                        h = All_file.objects.filter(username=username,
                                                  name_rk=name_rk)
                        h.update(mp=mp)
                    data['form1'] = ClearForm(initial={
                                       'name_rk': rk,
                                       'comments': comments,
                                       'landing': landing,
                                       'access': access
                                           })
            if request.method=='POST' and 'form2' in request.POST:
                form2 = ReportForm(request.POST, request.FILES)
                if form2.is_valid():
                    report = request.FILES.get('report')
                    a = All_file.objects.get(username=username,
                                                     name_rk=name_rk)
                    a.report = report
                    a.save()
                    n = f'media/{a.report.name}'
                    hol = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                    report_xl = pd.read_excel(os.path.join(hol, n).replace('\\', '/'), engine='openpyxl',
                                         header=None)
                    client = report_xl[0][0]
                    categ_cl = report[0][1]
                    count_site = ''
                    count_format = ''
                    count_ctr = ''
                    count_vtr = ''
                    count_lid = ''
                    for y in range(len(report_xl.values)):
                        if 'VTR,%' in report_xl.values[y] or 'VTR' in report_xl.values[y]:
                            count_head = y
                            for i in report_xl.values[y]: #i - string
                                if 'CTR' in i:
                                    for i in range(count_ctr, count_ctr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_ctr = i+1
                                        if '%' in report_xl[i]:
                                            count_ctr = i
                                if 'VTR' in i:
                                    for i in range(count_vtr, count_vtr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_vtr = i+1
                                        if '%' in report_xl[i]:
                                            count_vtr = i
                                if 'Площадка' in i or 'площадка' in i or 'Site' in i or 'Сайт' in i or 'Рекламная площадка' in i:
                                    count_site = i
                                if 'формат' in i or 'Format' in i or 'Формат рекламных материалов' in i or 'Формат' in i or 'Размер (в пикселях) / Формат' in i:
                                    count_format = i
                                if 'Звонки' in i or 'звонки' in i or 'CPL' in i or 'Заявки' in i or 'заявки' in i or 'Лиды/постклики' in i:
                                    count_lid = i
                        elif 'Площадка' in report_xl.values[y] or 'площадка' in report_xl.values[y] or 'Site' in report_xl.values[y] or 'Рекламная площадка' in report_xl.values[y]:
                            count_head = y
                            for i in report_xl.values[y]:
                                if 'CTR' in i:
                                    for i in range(count_ctr, count_ctr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_ctr = i+1
                                        if '%' in report_xl[i]:
                                            count_ctr = i
                                if 'VTR' in i:
                                    for i in range(count_vtr, count_vtr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_vtr = i+1
                                        if '%' in report_xl[i]:
                                            count_vtr = i
                                if 'Площадка' in i or 'площадка' in i or 'Site' in i or 'Сайт' in i or 'Рекламная площадка' in i:
                                    count_site = i
                                if 'формат' in i or 'Format' in i or 'Формат рекламных материалов' in i or 'Формат' in i or 'Размер (в пикселях) / Формат' in i:
                                    count_format = i
                                if 'Звонки' in i or 'звонки' in i or 'CPL' in i or 'Заявки' in i or 'заявки' in i or 'Лиды/постклики' in i:
                                    count_lid = i
                        elif 'формат' in report_xl.values[y] or 'Format' in report_xl.values[y] or 'Формат рекламных материалов' in report_xl.values[y] or 'Формат' in report_xl.values[y]:
                            count_head = y
                            for i in report_xl.values[y]:
                                if 'CTR' in i:
                                    for i in range(count_ctr, count_ctr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_ctr = i+1
                                        if '%' in report_xl[i]:
                                            count_ctr = i
                                if 'VTR' in i:
                                    for i in range(count_vtr, count_vtr+3):
                                        if 'Факт' in report_xl[i]:
                                            count_vtr = i+1
                                        if '%' in report_xl[i]:
                                            count_vtr = i
                                if 'Площадка' in i or 'площадка' in i or 'Site' in i or 'Сайт' in i or 'Рекламная площадка' in i:
                                    count_site = i
                                if 'формат' in i or 'Format' in i or 'Формат рекламных материалов' in i or 'Формат' in i or 'Размер (в пикселях) / Формат' in i:
                                    count_format = i
                                if 'Звонки' in i or 'звонки' in i or 'CPL' in i or 'Заявки' in i or 'заявки' in i or 'Лиды/постклики' in i:
                                    count_lid = i
                        if count_site != '' or count_format != '' or count_ctr != '' or count_vtr != '' or count_lid != '':
                            break

                    report_xl = pd.read_excel(os.path.join(hol, n).replace('\\', '/'), engine='openpyxl',
                                         header=None, skiprows=count_head)
                    if count_site!='':
                        len_list = len(report_xl[count_site])
                    elif count_format!='':
                        len_list = len(report_xl[count_format])
                    elif count_vtr!='':
                        len_list = len(report_xl[count_vtr])
                    elif count_ctr!='':
                        len_list = len(report_xl[count_ctr])
                    else:
                        len_list = 0
                    if len_list != 0:
                        try:
                            if count_vtr == '' and count_ctr != '' and count_lid != '':
                                p = pd.DataFrame([client]*len_list+[categ_cl]*len_list+report_xl[count_site]+report_xl[count_format]+['']*len_list+report_xl[count_ctr]+
                                    report_xl[count_lid])
                            elif count_ctr == '' and count_vtr != '' and count_lid != '':
                                p = pd.DataFrame([client]*len_list+[categ_cl]*len_list+report_xl[count_site]+report_xl[count_format]+report_xl[count_vtr]+['']*len_list+
                                    report_xl[count_lid])
                            else:
                                p = pd.DataFrame([client]*len_list+[categ_cl]*len_list+report_xl[count_site]+report_xl[count_format]+report_xl[count_vtr]+
                                    report_xl[count_vtr]+['']*len_list)

                            wb = openpyxl.load_workbook(filename=os.path.join(hol, f"media/clients/{translit(username)}/report.xlsx"))

                            w = wb.worksheets[0]
                            sheet = wb.active
                            for r in dataframe_to_rows(p, index=None, header=None):
                                w.append(r)
                            wb.save()
                        except:
                            pass
            return render(request, 'but_cleared.html', data)
        except MultipleObjectsReturned:
            pass
    else:
        return redirect('exel:login')



def utm(request, pk):
    if request.user.is_authenticated:
        username = request.user.username
        data = {
           'file': name_rk,
           'form': UtmForm(),

           }
        if request.method == 'POST':
            form = UtmForm(request.POST, request.FILES)
            if form.is_valid():
                u = request.FILES.get('utm')

                a = Cleared.objects.get(name_rk=name_rk)
                a.utm = u
                a.save()
                data['utm_name'] = u

        data['form'] = UtmForm()
        data['utm'] = Cleared.objects.get(name_rk=name_rk)
        return render(request, 'utm.html', data)
    else:
        return redirect('exel:login')

def materials(request, pk):
    if request.user.is_authenticated:
        a = Feed.objects.filter(name_rk=name_rk)[::-1]
        data = {
            'name_rk': name_rk,
           'files': [FeedFile.objects.filter(feed_id=i.pk) for i in a],
           'form': FileModelForm(),
           }
        cl = Cleared.objects.get(name_rk=name_rk)
        username = request.user.username
        if request.method == 'POST':
            file_form = FileModelForm(request.POST, request.FILES)
            files = request.FILES.getlist('file') #field name in model
            if file_form.is_valid():
                feed_instance = Feed.objects.create(name_rk=name_rk, client=cl.client,
                                                    username=username)
                for f in files:
                    file_instance = FeedFile(file=f, feed=feed_instance)
                    file_instance.save()
        data['form'] = FileModelForm()
        a = Feed.objects.filter(name_rk=name_rk)[::-1]
        data['files'] = [FeedFile.objects.filter(feed_id=i.pk) for i in a]
        return render(request, 'materials.html', data)
    else:
        return redirect('exel:login')

def complete(request, pk):
    if request.user.is_authenticated:
        username = request.user.username
        f = Brief.objects.filter(username=username,
                               name_rk=name_rk)[::-1][0]
        try:
            Complete.objects.get(username=username, name_rk=name_rk,
                                 client=f.client)
        except ObjectDoesNotExist:
            Complete.objects.create(username=username, name_rk=name_rk,
                                      client=f.client, budget=f.budget,
                                      period_c=f.period_c, period_p=f.period_p)
        except MultipleObjectsReturned:
            pass
        return main(request)
    else:
        return redirect('exel:login')
